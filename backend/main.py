from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
import json
import os

app = FastAPI()

# กำหนดค่า CORS
origins = [
    "http://localhost:5173",  # เพิ่ม URL ของ frontend ที่คุณใช้
    "http://127.0.0.1:5173",  # ถ้าคุณใช้ localhost
    "http://localhost:8080",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  # อนุญาตโดเมนเหล่านี้
    allow_credentials=True,
    allow_methods=["*"],  # อนุญาตวิธีการ HTTP ทั้งหมด
    allow_headers=["*"],  # อนุญาต headers ทั้งหมด
)

class JsonData(BaseModel):
    data: dict

output_directory = os.path.expanduser("~/Desktop/")
os.makedirs(output_directory, exist_ok=True)


@app.get("/")
async def read_root():
    return {"message": "Welcome to FastAPI-Vue Project"}



@app.post("/uploadfile/")
async def upload_file(file: UploadFile = File(...)):
    # Check File Format
    if not file or not file.filename:
        return JSONResponse(content={"error": "File format not supported. Please upload a .csv or .xlsx file."}, status_code=400)

    if not file.filename.endswith('.csv') and not file.filename.endswith('.xlsx'):
        return JSONResponse(content={"error": "File format not supported. Please upload a .csv or .xlsx file."}, status_code=400)

    # Read File
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file.file)

        elif file.filename.endswith('.xlsx'):
            df = pd.read_excel(file.file)

        # Convert Data to JSON
        json_data = json.loads(df.to_json(orient='split'))
        print(json_data)
        
        return JSONResponse(content={"data": json_data}, status_code=200)

    except pd.errors.EmptyDataError:
        return JSONResponse(content={"error": "No data in the file."}, status_code=400)
    except pd.errors.ParserError:
        return JSONResponse(content={"error": "Invalid file format."}, status_code=400)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.post("/downloadfile/")
async def download_file(json_data: JsonData):
    try:
        # Validate JSON data structure
        if not json_data or not json_data.data:
            raise HTTPException(status_code=400, detail="Invalid request body")

        # Check if 'index', 'columns', and 'data' keys exist in json_data
        required_keys = {'index', 'columns', 'data'}
        if not required_keys.issubset(json_data.data):
            raise HTTPException(status_code=400, detail="Missing 'index', 'columns', or 'data' in the request body")

        # Convert JSON to DataFrame
        df = pd.DataFrame(data=json_data.data['data'], columns=json_data.data['columns'])

        # Create Excel file in output directory
        os.makedirs(output_directory, exist_ok=True)  # Ensure output directory exists
        
        # Determine unique file name
        base_filename = "output.xlsx"
        file_path = os.path.join(output_directory, base_filename)
        counter = 1

        while os.path.exists(file_path):
            # Create a new file name with a counter
            file_path = os.path.join(output_directory, f"output_{counter}.xlsx")
            counter += 1
            
        df.to_excel(file_path, index=False)

        # Load workbook and apply styles
        workbook = load_workbook(file_path)
        sheet = workbook.active

        style_header(sheet)
        style_body(sheet)

        # Save styled workbook
        workbook.save(file_path)

        # Send Excel file to download
        return FileResponse(file_path, filename="output.xlsx", media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except KeyError as e:
        raise HTTPException(status_code=400, detail=f"Missing key: {str(e)}")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Value error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
       
def style_header(sheet):
    """Apply header styles to the sheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            
def style_body(sheet):
    """Apply body styles to the sheet."""
    body_font = Font(color="000000")
    body_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    empty_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = body_font
            # Check if cell is empty and apply gray fill
            if cell.value is None or cell.value == "":
                cell.fill = empty_fill
            else:
                if cell.column == 1:  # First column
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Same as header
                    cell.font = Font(bold=True, color="FFFFFF")  # Make it bold
                    cell.border = border
                else:
                    cell.fill = body_fill  # Apply light blue fill for non-empty cells
            cell.alignment = center_alignment
